#excel_writer.py
import os
from openpyxl import load_workbook
#openpyxl讀取寫入修改excel的函式

def append_to_excel(
    excel_file,
    second_area,
    date,
    doc_no,
    subject):
 # 安全防禦：如果 Excel 檔案根本不存在，直接告知失敗
    if not os.path.exists(excel_file):
        print(f"[錯誤] Excel 檔案不存在: {excel_file}")
        return False
    try:

        wb = load_workbook(excel_file)

        if second_area not in wb.sheetnames:

            print(f"找不到工作表：{second_area}")
            # 建立新工作表
            ws = wb.create_sheet(title=second_area)
            # 自動寫入第一列的欄位標題
            ws.append(["發文日期", "公文文號", "公文主旨"])
        else:
            ws = wb[second_area]

        # 檢查文號是否已存在，Excel 欄位順序是：A欄(日期)、B欄(文號)、C欄(主旨)，則 row[1] 代表 B 欄
        #min_row=2標題下一列開始記錄
        for row in ws.iter_rows(
            min_row=2,max_col=3,
            values_only=True):
            #防空值Excel底部是「完全空白的橫列，直接抓會報錯
            #row：確保這一橫列不是 None，len(row) > 1，至少有 2 欄以上的資料、row[1]B欄真有值
            if row and len(row) > 1 and row[1] is not None:
                cell_value = str(row[1]).strip()
                target_value = str(doc_no).strip()
                
                # 安全攔截：只有當兩個字串完全一模一樣、且都不是空字串時，才算重複
                if cell_value and cell_value == target_value:
                    print(f"[Excel跳過] 文號已存在於 [{second_area}] 工作表：{doc_no}")
                    wb.close()
                    return "EXIST"

                

        ws.append([
            date,
            doc_no,
            subject
        ])

        wb.save(excel_file)
        wb.close() # 務必關閉檔案，釋放系統資源
        print("成功寫入excel")
        return True

    except Exception as e:

        print(f"程式執行時，務必關閉EXCEL: {e}")

        return False